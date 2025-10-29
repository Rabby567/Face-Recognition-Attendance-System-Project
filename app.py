"""
Face Recognition Attendance System - Universal CCTV Support
Supports: RTSP, MJPEG, HTTP, DroidCam, IP Webcam, and all camera types
"""

from flask import Flask, render_template, request, jsonify, session, redirect, url_for, Response
from flask_cors import CORS
import cv2
import numpy as np
import csv
import os
from datetime import datetime, date
from pathlib import Path
import openpyxl
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font, PatternFill, Alignment
import pickle
import base64
import io
from PIL import Image
import secrets
import hashlib
import threading
import requests

app = Flask(__name__)
app.secret_key = secrets.token_hex(32)
CORS(app)

DEFAULT_USERNAME = "admin"
DEFAULT_PASSWORD = "admin"

camera_captures = {}
camera_locks = {}

class AttendanceSystem:
    def __init__(self):
        self.base_dir = Path.cwd()
        self.photos_dir = self.base_dir / "photos"
        self.attendance_dir = self.base_dir / "attendance"
        self.registration_file = self.base_dir / "registration.csv"
        self.attendance_cache_file = self.base_dir / "attendance_cache.pkl"
        self.settings_file = self.base_dir / "settings.pkl"
        self.admin_file = self.base_dir / "admin.pkl"
        
        self.photos_dir.mkdir(exist_ok=True)
        self.attendance_dir.mkdir(exist_ok=True)
        
        cascade_path = cv2.data.haarcascades + 'haarcascade_frontalface_default.xml'
        self.face_cascade = cv2.CascadeClassifier(cascade_path)
        
        self.known_face_data = []
        self.known_face_images = {}
        self.today_attended = self.load_attendance_cache()
        
        settings = self.load_settings()
        self.late_time = settings.get('late_time', '09:00')
        self.auto_start_time = settings.get('auto_start_time', '07:00')
        self.auto_end_time = settings.get('auto_end_time', '18:00')
        
        if not self.registration_file.exists():
            with open(self.registration_file, 'w', newline='', encoding='utf-8') as f:
                writer = csv.writer(f)
                writer.writerow(['Employee_ID', 'Name', 'Phone', 'Address', 'Photo_Path', 'Registration_Date'])
        
        self.load_employee_data()
    
    def load_attendance_cache(self):
        try:
            if self.attendance_cache_file.exists():
                with open(self.attendance_cache_file, 'rb') as f:
                    cache = pickle.load(f)
                    if cache.get('date') == str(date.today()):
                        return cache.get('attended', set())
        except:
            pass
        return set()
    
    def save_attendance_cache(self):
        try:
            cache = {'date': str(date.today()), 'attended': self.today_attended}
            with open(self.attendance_cache_file, 'wb') as f:
                pickle.dump(cache, f)
        except:
            pass
    
    def load_settings(self):
        try:
            if self.settings_file.exists():
                with open(self.settings_file, 'rb') as f:
                    return pickle.load(f)
        except:
            pass
        return {
            'late_time': '09:00', 
            'auto_start_time': '07:00', 
            'auto_end_time': '18:00',
            'cctv_cameras': []
        }
    
    def save_settings(self, settings):
        try:
            with open(self.settings_file, 'wb') as f:
                pickle.dump(settings, f)
            self.late_time = settings.get('late_time', self.late_time)
            self.auto_start_time = settings.get('auto_start_time', self.auto_start_time)
            self.auto_end_time = settings.get('auto_end_time', self.auto_end_time)
            return True
        except:
            return False
    
    def get_cctv_cameras(self):
        settings = self.load_settings()
        return settings.get('cctv_cameras', [])
    
    def add_cctv_camera(self, name, url):
        settings = self.load_settings()
        cameras = settings.get('cctv_cameras', [])
        
        for cam in cameras:
            if cam['name'] == name or cam['url'] == url:
                return False
        
        camera_id = len(cameras) + 1
        cameras.append({
            'id': camera_id,
            'name': name,
            'url': url
        })
        
        settings['cctv_cameras'] = cameras
        self.save_settings(settings)
        return True
    
    def remove_cctv_camera(self, camera_id):
        settings = self.load_settings()
        cameras = settings.get('cctv_cameras', [])
        
        if camera_id in camera_captures:
            try:
                camera_captures[camera_id].release()
                del camera_captures[camera_id]
            except:
                pass
        
        cameras = [cam for cam in cameras if cam['id'] != camera_id]
        
        settings['cctv_cameras'] = cameras
        self.save_settings(settings)
        return True
    
    def hash_password(self, password):
        return hashlib.sha256(password.encode()).hexdigest()
    
    def load_admin(self):
        try:
            if self.admin_file.exists():
                with open(self.admin_file, 'rb') as f:
                    return pickle.load(f)
        except:
            pass
        return None
    
    def save_admin(self, username, password):
        try:
            admin_data = {
                'username': username,
                'password_hash': self.hash_password(password)
            }
            with open(self.admin_file, 'wb') as f:
                pickle.dump(admin_data, f)
            return True
        except:
            return False
    
    def verify_admin(self, username, password):
        admin = self.load_admin()
        
        if admin is None:
            if username == DEFAULT_USERNAME and password == DEFAULT_PASSWORD:
                return True, 'default'
            return False, None
        
        if admin['username'] == username and admin['password_hash'] == self.hash_password(password):
            return True, 'custom'
        
        return False, None
    
    def load_employee_data(self):
        self.known_face_data = []
        self.known_face_images = {}
        
        if self.registration_file.exists():
            try:
                with open(self.registration_file, 'r', encoding='utf-8') as f:
                    reader = csv.DictReader(f)
                    for row in reader:
                        self.known_face_data.append(row)
                        
                        photo_path = row.get('Photo_Path', '')
                        if os.path.exists(photo_path):
                            img = cv2.imread(photo_path, cv2.IMREAD_GRAYSCALE)
                            if img is not None:
                                faces = self.face_cascade.detectMultiScale(img, 1.1, 4)
                                if len(faces) > 0:
                                    x, y, w, h = faces[0]
                                    face_roi = img[y:y+h, x:x+w]
                                    face_resized = cv2.resize(face_roi, (100, 100))
                                    self.known_face_images[row['Employee_ID']] = face_resized
            except:
                pass
    
    def register_employee(self, emp_id, name, phone, address, image_data):
        try:
            for emp in self.known_face_data:
                if emp['Employee_ID'] == emp_id:
                    return {'success': False, 'message': f'Employee ID {emp_id} already exists'}
            
            try:
                if ',' in image_data:
                    image_bytes = base64.b64decode(image_data.split(',')[1])
                else:
                    image_bytes = base64.b64decode(image_data)
                    
                image = Image.open(io.BytesIO(image_bytes))
                image_np = np.array(image)
                
                if len(image_np.shape) == 2:
                    image_bgr = cv2.cvtColor(image_np, cv2.COLOR_GRAY2BGR)
                elif image_np.shape[2] == 4:
                    image_bgr = cv2.cvtColor(image_np, cv2.COLOR_RGBA2BGR)
                elif image_np.shape[2] == 3:
                    image_bgr = cv2.cvtColor(image_np, cv2.COLOR_RGB2BGR)
                else:
                    image_bgr = image_np
            except Exception as e:
                return {'success': False, 'message': f'Invalid image: {str(e)}'}
            
            gray = cv2.cvtColor(image_bgr, cv2.COLOR_BGR2GRAY)
            faces = self.face_cascade.detectMultiScale(gray, 1.1, 4)
            
            if len(faces) == 0:
                return {'success': False, 'message': 'No face detected'}
            
            photo_filename = f"{emp_id}_{name.replace(' ', '_')}.jpg"
            photo_path = self.photos_dir / photo_filename
            cv2.imwrite(str(photo_path), image_bgr)
            
            x, y, w, h = faces[0]
            face_roi = gray[y:y+h, x:x+w]
            face_resized = cv2.resize(face_roi, (100, 100))
            self.known_face_images[emp_id] = face_resized
            
            with open(self.registration_file, 'a', newline='', encoding='utf-8') as f:
                writer = csv.writer(f)
                writer.writerow([emp_id, name, phone, address, str(photo_path), datetime.now().strftime('%Y-%m-%d %H:%M:%S')])
            
            self.known_face_data.append({
                'Employee_ID': emp_id,
                'Name': name,
                'Phone': phone,
                'Address': address,
                'Photo_Path': str(photo_path)
            })
            
            return {'success': True, 'message': 'Registration successful', 'emp_id': emp_id, 'name': name}
        except Exception as e:
            return {'success': False, 'message': str(e)}
    
    def mark_attendance(self, emp_data, timestamp):
        try:
            today = date.today()
            year_month = today.strftime('%Y-%m')
            month_dir = self.attendance_dir / year_month
            month_dir.mkdir(exist_ok=True)
            
            excel_file = month_dir / f"attendance_{today.strftime('%Y-%m-%d')}.xlsx"
            
            if excel_file.exists():
                wb = load_workbook(excel_file)
            else:
                wb = Workbook()
                wb.remove(wb.active)
            
            sheet_name = f"{emp_data['Employee_ID']}_{emp_data['Name'][:20]}"
            if sheet_name in wb.sheetnames:
                ws = wb[sheet_name]
            else:
                ws = wb.create_sheet(sheet_name)
                headers = ['Date', 'Time', 'Employee ID', 'Name', 'Status', 'Minutes Late']
                ws.append(headers)
                for cell in ws[1]:
                    cell.font = Font(bold=True, color="FFFFFF")
                    cell.fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
                    cell.alignment = Alignment(horizontal="center")
            
            arrival_time = timestamp.time()
            late_time_obj = datetime.strptime(self.late_time, '%H:%M').time()
            
            status = 'Present'
            minutes_late = 0
            
            if arrival_time > late_time_obj:
                arrival_dt = datetime.combine(today, arrival_time)
                late_dt = datetime.combine(today, late_time_obj)
                minutes_late = int((arrival_dt - late_dt).total_seconds() / 60)
                status = 'Late'
            
            ws.append([
                today.strftime('%Y-%m-%d'),
                timestamp.strftime('%H:%M:%S'),
                emp_data['Employee_ID'],
                emp_data['Name'],
                status,
                minutes_late if minutes_late > 0 else ''
            ])
            
            last_row = ws.max_row
            if status == 'Late':
                for cell in ws[last_row]:
                    cell.fill = PatternFill(start_color="FFF3CD", end_color="FFF3CD", fill_type="solid")
            
            for column in ws.columns:
                max_length = 0
                column = list(column)
                for cell in column:
                    try:
                        if len(str(cell.value)) > max_length:
                            max_length = len(cell.value)
                    except:
                        pass
                adjusted_width = (max_length + 2)
                ws.column_dimensions[column[0].column_letter].width = adjusted_width
            
            wb.save(excel_file)
            return str(excel_file), status, minutes_late
        except:
            return None, 'Error', 0
    
    def compare_faces(self, face1, face2):
        try:
            if face1.shape != face2.shape:
                face2 = cv2.resize(face2, (face1.shape[1], face1.shape[0]))
            
            methods = [cv2.TM_CCOEFF_NORMED, cv2.TM_CCORR_NORMED]
            scores = []
            
            for method in methods:
                result = cv2.matchTemplate(face1, face2, method)
                scores.append(result[0][0])
            
            return np.mean(scores)
        except:
            return 0.0
    
    def process_frame(self, frame):
        results = []
        
        try:
            gray = cv2.cvtColor(frame, cv2.COLOR_BGR2GRAY)
            faces = self.face_cascade.detectMultiScale(gray, 1.1, 4, minSize=(50, 50))
            
            for (x, y, w, h) in faces:
                face_roi = gray[y:y+h, x:x+w]
                face_resized = cv2.resize(face_roi, (100, 100))
                
                name = "Unknown"
                emp_id = None
                attended = False
                best_match_score = 0
                best_match_emp = None
                
                for emp_data in self.known_face_data:
                    emp_id_check = emp_data['Employee_ID']
                    
                    if emp_id_check in self.known_face_images:
                        stored_face = self.known_face_images[emp_id_check]
                        score = self.compare_faces(stored_face, face_resized)
                        
                        if score > best_match_score:
                            best_match_score = score
                            best_match_emp = emp_data
                
                if best_match_score > 0.65:
                    name = best_match_emp['Name']
                    emp_id = best_match_emp['Employee_ID']
                    
                    if emp_id not in self.today_attended:
                        timestamp = datetime.now()
                        self.mark_attendance(best_match_emp, timestamp)
                        self.today_attended.add(emp_id)
                        self.save_attendance_cache()
                        attended = True
                
                results.append({
                    'name': name,
                    'emp_id': emp_id,
                    'box': {'top': y, 'right': x+w, 'bottom': y+h, 'left': x},
                    'attended': attended,
                    'already_attended': emp_id in self.today_attended if emp_id else False
                })
        except:
            pass
        
        return results

system = AttendanceSystem()

@app.route('/')
def index():
    return render_template('index.html')

@app.route('/dashboard')
def dashboard():
    if 'admin_logged_in' not in session:
        return redirect(url_for('login'))
    return render_template('dashboard.html')

@app.route('/login')
def login():
    return render_template('login.html')

@app.route('/api/login', methods=['POST'])
def api_login():
    try:
        data = request.json
        username = data.get('username', '')
        password = data.get('password', '')
        
        is_valid, auth_type = system.verify_admin(username, password)
        
        if is_valid:
            session['admin_logged_in'] = True
            session['username'] = username
            session['auth_type'] = auth_type
            return jsonify({'success': True, 'auth_type': auth_type})
        else:
            return jsonify({'success': False, 'message': 'Invalid credentials'})
    except:
        return jsonify({'success': False, 'message': 'Login failed'})

@app.route('/api/logout', methods=['POST'])
def logout():
    session.clear()
    return jsonify({'success': True})

@app.route('/api/change-password', methods=['POST'])
def change_password():
    if 'admin_logged_in' not in session:
        return jsonify({'success': False, 'message': 'Unauthorized'}), 401
    
    try:
        data = request.json
        old_password = data.get('old_password', '')
        new_username = data.get('username', '')
        new_password = data.get('password', '')
        
        if not old_password or not new_username or not new_password:
            return jsonify({'success': False, 'message': 'All fields required'})
        
        current_username = session.get('username')
        is_valid, _ = system.verify_admin(current_username, old_password)
        
        if not is_valid:
            return jsonify({'success': False, 'message': 'Incorrect old password'})
        
        if system.save_admin(new_username, new_password):
            session.clear()
            return jsonify({'success': True, 'message': 'Credentials updated'})
        else:
            return jsonify({'success': False, 'message': 'Failed to save'})
    except:
        return jsonify({'success': False, 'message': 'Update failed'})

@app.route('/api/register', methods=['POST'])
def register():
    if 'admin_logged_in' not in session:
        return jsonify({'success': False, 'message': 'Unauthorized'}), 401
    
    try:
        data = request.json
        result = system.register_employee(
            data.get('emp_id', ''),
            data.get('name', ''),
            data.get('phone', ''),
            data.get('address', ''),
            data.get('image', '')
        )
        return jsonify(result)
    except:
        return jsonify({'success': False, 'message': 'Registration failed'}), 500

@app.route('/api/employees', methods=['GET'])
def get_employees():
    if 'admin_logged_in' not in session:
        return jsonify({'employees': [], 'count': 0}), 401
    
    employees = system.known_face_data
    return jsonify({'employees': employees, 'count': len(employees)})

@app.route('/api/process-frame', methods=['POST'])
def process_frame():
    try:
        data = request.json
        image_data = data['frame']
        
        if ',' in image_data:
            image_bytes = base64.b64decode(image_data.split(',')[1])
        else:
            image_bytes = base64.b64decode(image_data)
            
        image = Image.open(io.BytesIO(image_bytes))
        frame = np.array(image)
        
        if len(frame.shape) == 3 and frame.shape[2] == 3:
            frame_bgr = cv2.cvtColor(frame, cv2.COLOR_RGB2BGR)
        else:
            frame_bgr = frame
        
        results = system.process_frame(frame_bgr)
        return jsonify({'success': True, 'faces': results})
    except:
        return jsonify({'success': False, 'message': 'Processing failed'})

@app.route('/api/attendance-today', methods=['GET'])
def attendance_today():
    if 'admin_logged_in' not in session:
        return jsonify({'count': 0, 'employees': []}), 401
    
    attended_ids = list(system.today_attended)
    attended_employees = []
    
    for emp_data in system.known_face_data:
        if emp_data['Employee_ID'] in attended_ids:
            attended_employees.append(emp_data)
    
    return jsonify({
        'count': len(attended_employees),
        'employees': attended_employees,
        'date': str(date.today())
    })

@app.route('/api/stats', methods=['GET'])
def get_stats():
    total_employees = len(system.known_face_data)
    today_attendance = len(system.today_attended)
    attendance_rate = (today_attendance / total_employees * 100) if total_employees > 0 else 0
    
    current_time = datetime.now().time()
    start_time = datetime.strptime(system.auto_start_time, '%H:%M').time()
    end_time = datetime.strptime(system.auto_end_time, '%H:%M').time()
    within_hours = start_time <= current_time <= end_time
    
    return jsonify({
        'total_employees': total_employees,
        'today_attendance': today_attendance,
        'attendance_rate': round(attendance_rate, 1),
        'date': str(date.today()),
        'within_hours': within_hours,
        'operating_hours': f"{system.auto_start_time} - {system.auto_end_time}"
    })

@app.route('/api/download-csv', methods=['POST'])
def download_csv():
    if 'admin_logged_in' not in session:
        return jsonify({'success': False, 'message': 'Unauthorized'}), 401
    
    try:
        data = request.json
        download_type = data.get('type', 'day')
        selected_date = data.get('date', str(date.today()))
        
        try:
            dt = datetime.strptime(selected_date, '%Y-%m-%d')
        except:
            dt = date.today()
        
        csv_data = []
        csv_data.append(['Date', 'Time', 'Employee ID', 'Name', 'Phone', 'Status', 'Minutes Late'])
        
        if download_type == 'day':
            year_month = dt.strftime('%Y-%m')
            month_dir = system.attendance_dir / year_month
            excel_file = month_dir / f"attendance_{dt.strftime('%Y-%m-%d')}.xlsx"
            
            if excel_file.exists():
                wb = load_workbook(excel_file)
                for sheet_name in wb.sheetnames:
                    ws = wb[sheet_name]
                    for row in ws.iter_rows(min_row=2, values_only=True):
                        if row[0]:
                            emp_id = row[2]
                            phone = ''
                            for emp in system.known_face_data:
                                if emp['Employee_ID'] == emp_id:
                                    phone = emp.get('Phone', '')
                                    break
                            minutes_late = row[5] if len(row) > 5 else ''
                            csv_data.append([row[0], row[1], row[2], row[3], phone, row[4], minutes_late])
        
        elif download_type == 'month':
            year_month = dt.strftime('%Y-%m')
            month_dir = system.attendance_dir / year_month
            
            if month_dir.exists():
                for excel_file in sorted(month_dir.glob('attendance_*.xlsx')):
                    wb = load_workbook(excel_file)
                    for sheet_name in wb.sheetnames:
                        ws = wb[sheet_name]
                        for row in ws.iter_rows(min_row=2, values_only=True):
                            if row[0]:
                                emp_id = row[2]
                                phone = ''
                                for emp in system.known_face_data:
                                    if emp['Employee_ID'] == emp_id:
                                        phone = emp.get('Phone', '')
                                        break
                                minutes_late = row[5] if len(row) > 5 else ''
                                csv_data.append([row[0], row[1], row[2], row[3], phone, row[4], minutes_late])
        
        elif download_type == 'year':
            year = dt.strftime('%Y')
            for month_dir in sorted(system.attendance_dir.glob(f'{year}-*')):
                for excel_file in sorted(month_dir.glob('attendance_*.xlsx')):
                    wb = load_workbook(excel_file)
                    for sheet_name in wb.sheetnames:
                        ws = wb[sheet_name]
                        for row in ws.iter_rows(min_row=2, values_only=True):
                            if row[0]:
                                emp_id = row[2]
                                phone = ''
                                for emp in system.known_face_data:
                                    if emp['Employee_ID'] == emp_id:
                                        phone = emp.get('Phone', '')
                                        break
                                minutes_late = row[5] if len(row) > 5 else ''
                                csv_data.append([row[0], row[1], row[2], row[3], phone, row[4], minutes_late])
        
        import io as csv_io
        output = csv_io.StringIO()
        writer = csv.writer(output)
        writer.writerows(csv_data)
        csv_string = output.getvalue()
        
        if download_type == 'day':
            filename = f"attendance_{dt.strftime('%Y-%m-%d')}.csv"
        elif download_type == 'month':
            filename = f"attendance_{dt.strftime('%Y-%m')}.csv"
        else:
            filename = f"attendance_{dt.strftime('%Y')}.csv"
        
        return jsonify({
            'success': True,
            'csv': csv_string,
            'filename': filename,
            'records': len(csv_data) - 1
        })
    except:
        return jsonify({'success': False, 'message': 'Download failed'})

@app.route('/api/settings', methods=['GET'])
def get_settings():
    if 'admin_logged_in' not in session:
        return jsonify({'success': False, 'message': 'Unauthorized'}), 401
    
    return jsonify({
        'success': True,
        'late_time': system.late_time,
        'auto_start_time': system.auto_start_time,
        'auto_end_time': system.auto_end_time
    })

@app.route('/api/settings', methods=['POST'])
def update_settings():
    if 'admin_logged_in' not in session:
        return jsonify({'success': False, 'message': 'Unauthorized'}), 401
    
    try:
        data = request.json
        current_settings = system.load_settings()
        
        settings = {
            'late_time': data.get('late_time', system.late_time),
            'auto_start_time': data.get('auto_start_time', system.auto_start_time),
            'auto_end_time': data.get('auto_end_time', system.auto_end_time),
            'cctv_cameras': current_settings.get('cctv_cameras', [])
        }
        
        if system.save_settings(settings):
            return jsonify({
                'success': True,
                'message': 'Settings updated',
                'settings': settings
            })
        else:
            return jsonify({'success': False, 'message': 'Failed to save'})
    except:
        return jsonify({'success': False, 'message': 'Update failed'})

@app.route('/api/admin-info', methods=['GET'])
def get_admin_info():
    if 'admin_logged_in' not in session:
        return jsonify({'logged_in': False}), 401
    
    return jsonify({
        'logged_in': True,
        'username': session.get('username'),
        'auth_type': session.get('auth_type')
    })

@app.route('/api/cctv-cameras', methods=['GET'])
def get_cctv_cameras():
    if 'admin_logged_in' not in session:
        return jsonify({'success': False, 'message': 'Unauthorized'}), 401
    
    cameras = system.get_cctv_cameras()
    return jsonify({'success': True, 'cameras': cameras})

@app.route('/api/public/cctv-cameras', methods=['GET'])
def get_public_cctv_cameras():
    cameras = system.get_cctv_cameras()
    return jsonify({'success': True, 'cameras': cameras})

@app.route('/api/cctv-cameras', methods=['POST'])
def add_cctv_camera():
    if 'admin_logged_in' not in session:
        return jsonify({'success': False, 'message': 'Unauthorized'}), 401
    
    try:
        data = request.json
        name = data.get('name', '')
        url = data.get('url', '')
        
        if not name or not url:
            return jsonify({'success': False, 'message': 'Name and URL required'})
        
        if system.add_cctv_camera(name, url):
            return jsonify({'success': True, 'message': 'Camera added'})
        else:
            return jsonify({'success': False, 'message': 'Camera already exists'})
    except:
        return jsonify({'success': False, 'message': 'Failed to add camera'})

@app.route('/api/cctv-cameras/<int:camera_id>', methods=['DELETE'])
def remove_cctv_camera(camera_id):
    if 'admin_logged_in' not in session:
        return jsonify({'success': False, 'message': 'Unauthorized'}), 401
    
    try:
        system.remove_cctv_camera(camera_id)
        return jsonify({'success': True, 'message': 'Camera removed'})
    except:
        return jsonify({'success': False, 'message': 'Failed to remove camera'})

@app.route('/api/cctv-stream/<int:camera_id>')
def cctv_stream(camera_id):
    cameras = system.get_cctv_cameras()
    camera = next((cam for cam in cameras if cam['id'] == camera_id), None)
    
    if not camera:
        return jsonify({'success': False, 'message': 'Camera not found'}), 404
    
    def generate_frames():
        camera_url = camera['url']
        
        try:
            if any(x in camera_url.lower() for x in ['mjpeg', 'video', ':4747', ':8080', 'droidcam']):
                print(f"Using MJPEG mode for: {camera_url}")
                response = requests.get(camera_url, stream=True, timeout=5)
                
                if response.status_code == 200:
                    bytes_data = b''
                    for chunk in response.iter_content(chunk_size=4096):
                        if not chunk:
                            continue
                        
                        bytes_data += chunk
                        
                        while True:
                            a = bytes_data.find(b'\xff\xd8')
                            b = bytes_data.find(b'\xff\xd9')
                            
                            if a != -1 and b != -1 and b > a:
                                jpg = bytes_data[a:b+2]
                                bytes_data = bytes_data[b+2:]
                                
                                yield (b'--frame\r\n'
                                       b'Content-Type: image/jpeg\r\n\r\n' + jpg + b'\r\n')
                            else:
                                break
            else:
                print(f"Using OpenCV mode for: {camera_url}")
                cap = cv2.VideoCapture(camera_url)
                cap.set(cv2.CAP_PROP_BUFFERSIZE, 1)
                
                if not cap.isOpened():
                    print(f"Failed to open camera: {camera_url}")
                    return
                
                while True:
                    success, frame = cap.read()
                    if not success:
                        print("Failed to read frame")
                        break
                    
                    frame = cv2.resize(frame, (640, 480))
                    
                    ret, buffer = cv2.imencode('.jpg', frame, [cv2.IMWRITE_JPEG_QUALITY, 80])
                    if not ret:
                        continue
                    
                    frame_bytes = buffer.tobytes()
                    
                    yield (b'--frame\r\n'
                           b'Content-Type: image/jpeg\r\n\r\n' + frame_bytes + b'\r\n')
                
                cap.release()
                
        except Exception as e:
            print(f"Stream error for camera {camera_id}: {str(e)}")
            import traceback
            traceback.print_exc()
    
    return Response(generate_frames(), mimetype='multipart/x-mixed-replace; boundary=frame')

@app.route('/api/public/cctv-stream/<int:camera_id>')
def public_cctv_stream(camera_id):
    return cctv_stream(camera_id)

if __name__ == '__main__':
    templates_dir = Path('templates')
    templates_dir.mkdir(exist_ok=True)
    
    print("="*60)
    print("Face Recognition Attendance System - Universal CCTV Support")
    print("="*60)
    print(f"Photos: {system.photos_dir}")
    print(f"Attendance: {system.attendance_dir}")
    print(f"Employees: {len(system.known_face_data)}")
    print(f"Default Login: admin/admin")
    print(f"Operating Hours: {system.auto_start_time} - {system.auto_end_time}")
    print("="*60)
    print("Supported Cameras:")
    print("  ✓ RTSP Streams (IP Cameras)")
    print("  ✓ MJPEG Streams (HTTP)")
    print("  ✓ DroidCam (Phone Camera)")
    print("  ✓ IP Webcam")
    print("  ✓ All Standard CCTV Cameras")
    print("="*60)
    print("Server: http://localhost:5000")
    print("="*60)
    
    app.run(host='0.0.0.0', port=5000, debug=True, threaded=True)